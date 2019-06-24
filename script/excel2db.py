from app.models import Staff, db, Record
from openpyxl import load_workbook

def init_staff_table():
    excel = load_workbook('new_staff.xlsx')
    # 获取sheet：
    table = excel.get_sheet_by_name('Sheet1')  # 通过表名获取
    # 获取行数和列数：
    rows = table.max_row  # 获取行数
    cols = table.max_column  # 获取列数
    # 获取单元格值：
    for row in range(2,rows+1):
        number = table.cell(row=row, column=1).value
        name = table.cell(row=row, column=2).value
        branch = table.cell(row=row, column=4).value
        duty = table.cell(row=row, column=5).value
        db.session.add(Staff(number=number,
                             name=name,
                             branch=branch,
                             duty=duty))
    db.session.commit()
    print('已在表staff添加{}条数据'.format(rows-1))

def clear_staff_table():
    for staff in Staff.query.all():
        db.session.delete(staff)
    db.session.commit()
    print('已删除staff表内所有数据')

def clear_record_stable():
    for record in Record.query.all():
        db.session.delete(record)
    db.session.commit()
    print('已删除record表内所有数据')

def init_record_stable():
    db.session.add(Record(name='jiage', number='sss', device=1))
    db.session.commit()

if __name__=="__main__":
    from app import create_app
    with create_app().app_context():
        try:
            # clear_staff_table()
            init_record_stable()
        except Exception as e:
            print('sql error:', e)




